import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from pandas import Timestamp

from cruds.earning_cruds import earnings_cruds
from helpers.enums.currency_enum import CurrencyEnum
from helpers.enums.inline_buttons_helper_enum import InlineButtonsHelperEnum
from helpers.enums.xlsx_enum import XlsxEnum

earnings_columns = ['ID', 'Time Created', 'Summa', 'Comment', 'Currency', 'Source Name', 'Admin Username', "Source"]

currency_mapper = {CurrencyEnum.EURO: "€", CurrencyEnum.UAH: "UAH", CurrencyEnum.DOLLAR: '$'}


def generate_xlsx_file(partial=False):
    earnings = pd.DataFrame(earnings_cruds.get_all_for_xlsx(partial), columns=earnings_columns)

    workbook = Workbook()

    sheet1 = workbook.create_sheet(title='Loans')
    _adapt_sheet_1(sheet1)

    _create_rows(sheet1, earnings)

    _wrap_text(sheet1, earnings_columns)

    output_file = XlsxEnum.NAME
    workbook.remove(workbook['Sheet'])

    _conditional_formatting(sheet1, earnings)

    earnings_range = f"C1:{get_column_letter(len(earnings.columns))}{len(earnings) + 1}"
    sheet1.auto_filter.ref = earnings_range

    workbook.save(output_file)

    return output_file


def _adapt_sheet_1(sheet):
    sheet.column_dimensions['A'].width = 15
    sheet.row_dimensions[1].height = 20

    sheet.column_dimensions['B'].width = 15
    sheet.row_dimensions[1].height = 15

    sheet.column_dimensions['C'].width = 15
    sheet.row_dimensions[1].height = 25

    sheet.column_dimensions['D'].width = 50
    sheet.row_dimensions[1].height = 90

    sheet.column_dimensions['E'].width = 15
    sheet.row_dimensions[1].height = 25

    sheet.column_dimensions['F'].width = 15
    sheet.row_dimensions[1].height = 25

    sheet.column_dimensions['G'].width = 15
    sheet.row_dimensions[1].height = 20

    sheet.column_dimensions['H'].width = 15
    sheet.row_dimensions[1].height = 20

    sheet.sheet_view.zoomScale = 130


def _create_rows(sheet, table):
    for r, row in enumerate([table.columns] + table.values.tolist(), start=1):
        for c, value in enumerate(row, start=1):
            if table.columns[c - 1] == 'Summa':  # Check if the column is 'Summa'
                if isinstance(value, str) and value not in earnings_columns:
                    value = float(value)
            elif table.columns[c - 1] == 'Time Created':
                if isinstance(value, Timestamp) and value not in earnings_columns:
                    value = pd.Timestamp(value).date()
            elif table.columns[c - 1] == 'Source Name':
                if isinstance(value, str) and value not in earnings_columns:
                    if value == InlineButtonsHelperEnum.OTHER:
                        value = ''
            elif table.columns[c - 1] == 'Currency':
                if isinstance(value, str) and value not in earnings_columns:
                    if value in [currency.value for currency in CurrencyEnum]:
                        value = currency_mapper.get(value)
            sheet.cell(row=r, column=c, value=value.replace('\\', '') if isinstance(value, str) else value)


def _conditional_formatting(sheet, model):
    for row in sheet.iter_rows(min_row=2, max_row=len(model) + 1):
        summa_cell = row[earnings_columns.index('Summa')]
        withdraw_cell = row[earnings_columns.index('Source')]

        withdraw_value = withdraw_cell.value.lower()
        summa_value = summa_cell.value

        if withdraw_value == 'withdraw':
            withdraw_cell.value = withdraw_value.upper()
            withdraw_cell.font = Font(color="FF0000")

        summa_cell.font = Font(color="C0504E" if summa_value < 0 else "4F6328")


def _wrap_text(sheet, columns):
    for column in columns:
        if column != 'A':
            column_letter = chr(ord('B') + columns.index(column))
            for cell in sheet[column_letter]:
                cell.alignment = cell.alignment.copy(wrap_text=True)
